"""Merge the 3-level module list into a 2-level functional description.

Input:  C:/Users/ron37/Desktop/体彩衍生品商城模块清单.xlsx
Output: C:/Users/ron37/Desktop/体彩衍生品商城模块清单_二级版.xlsx

- Group rows by (主模块, 一级模块, 二级模块)
- Merge 三级模块 items into a flowing 功能说明 paragraph
- Sum 人天 and 金额
- Renumber 编号
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:/Users/ron37/Desktop/体彩衍生品商城模块清单.xlsx"
DST = r"C:/Users/ron37/Desktop/体彩衍生品商城模块清单_二级版.xlsx"

# ============================================================
# 二级模块的功能说明模板（按 (主, 一, 二) 精准匹配）
# 模板中包含对三级功能点的整合文字。
# ============================================================
DESCRIPTIONS = {
    # ---------------- C 端小程序 ----------------
    ("C端小程序", "首页", "Banner 轮播"):
        "负责小程序首页 Banner 轮播图的整体展示与交互。支持自动与手动切换两种播放模式，"
        "点击 Banner 可跳转至对应商品详情或营销活动页；同时按运营后台下发的配置拉取"
        "并展示当前生效的营销活动信息。",

    ("C端小程序", "首页", "分类入口"):
        "首页金刚区分类入口。采用 4-8 个宫格布局呈现一级分类，每个宫格渲染对应分类的"
        "图标与名称；用户点击任一分类即可跳转至该分类下的商品列表页。",

    ("C端小程序", "首页", "热门商品"):
        "首页热门商品瀑布流区域。以卡片形式展示商品主图、名称、价格等核心信息，"
        "支持加载更多与分页机制；同时提供快捷加入购物车入口，便于用户在不进入"
        "详情页的情况下完成加购。",

    ("C端小程序", "首页", "Tabbar 导航"):
        "小程序底部 Tabbar 主导航。包含首页、分类、购物车、我的 4 个主 Tab，"
        "负责页面间的一级切换；并在有新消息或待办时显示未读小红点提醒。",

    ("C端小程序", "商品", "分类页"):
        "商品分类页。采用左侧一级分类 + 右侧二级分类的双栏布局，用户点击不同分类"
        "可即时刷新右侧商品列表，并支持分类间联动筛选与商品数据回填。",

    ("C端小程序", "商品", "商品列表"):
        "商品列表页。支持列表与网格两种视图切换，提供价格、销量、综合等维度的"
        "排序能力，并可按分类、标签、价格区间进行多维筛选，配合上下拉分页"
        "完成海量商品的高效浏览。",

    ("C端小程序", "商品", "商品详情"):
        "商品详情页核心模块。顶部展示商品图片轮播，提供 SKU 规格选择弹层（支持"
        "多维规格组合联动），实时进行数量加减与库存校验，富文本详情渲染保证"
        "运营富媒体内容正常展示；底部提供「加入购物车」与「立即购买」两个"
        "核心转化入口。",

    ("C端小程序", "购物车", "购物车列表"):
        "购物车列表页。按店铺或供货商进行商品分组渲染，支持单条选中与一键全选，"
        "数量增减实时同步，并自动识别并清理已失效或下架的商品。",

    ("C端小程序", "购物车", "购物车结算"):
        "购物车结算区。实时计算选中商品的总价合计，结算栏采用置底固定布局，"
        "始终保持可见；点击「去结算」可跳转至订单确认页完成下单。",

    ("C端小程序", "订单", "订单确认"):
        "订单确认页。默认带入已选收货地址，支持临时切换其他地址；展示待结算的"
        "商品清单与价格明细；提供配送方式与优惠券选择能力；底部展示费用合计"
        "并支持一键提交订单。",

    ("C端小程序", "订单", "支付页"):
        "订单支付页。负责唤起微信支付 SDK 完成支付，支付过程中实时轮询支付结果，"
        "并对支付失败、用户取消等异常场景进行统一处理与重试引导。",

    ("C端小程序", "订单", "订单列表"):
        "订单列表页。按状态（待付款/待发货/待收货/已完成/退款中）提供 Tab 切换，"
        "列表中展示订单卡片摘要信息；用户可在卡片上直接进行取消订单、付款、"
        "申请退款、确认收货等快捷操作。",

    ("C端小程序", "订单", "订单详情"):
        "订单详情页。展示订单从下单到完成的全生命周期状态流转时间线，"
        "并接入物流轨迹信息；同时提供售后服务入口，便于用户发起退款/售后。",

    ("C端小程序", "订单", "退款申请"):
        "退款申请流程。提供结构化的退款原因选择，支持填写退款金额与上传"
        "凭证图片；申请提交后可实时查询退款进度，并以退款状态机 UI "
        "清晰呈现各阶段（待审核/已同意/已拒绝/已退款）的视觉反馈。",

    ("C端小程序", "用户", "登录/注册"):
        "用户登录与注册模块。支持微信一键授权登录以及手机号 + 验证码的传统"
        "登录方式；登录态采用 Token 机制进行持久化存储与自动刷新，"
        "确保会话稳定不掉线。",

    ("C端小程序", "用户", "个人中心"):
        "个人中心页。顶部展示用户头像与基本信息卡片；下方聚合订单、收藏、"
        "优惠券等高频功能入口；底部提供设置与退出登录能力。",

    ("C端小程序", "用户", "收货地址"):
        "收货地址管理。支持地址列表查看、新增与编辑地址、设置默认地址；"
        "地址录入采用省市区三级联动选择器；同时提供地址簿导入与"
        "基础字段校验能力，保证地址数据完整可用。",

    # ---------------- 后台管理系统 ----------------
    ("后台管理系统", "系统认证", "登录"):
        "后台登录入口。支持账号 + 密码登录，集成图形验证码防刷，"
        "并提供「记住登录」与 Token 自动续签机制，确保管理员会话稳定。",

    ("后台管理系统", "系统认证", "RBAC 权限"):
        "RBAC 角色权限体系。提供角色与菜单的关联管理，根据当前用户角色"
        "动态渲染左侧菜单；并对页面级、按钮级操作进行细粒度权限校验，"
        "保证越权操作不可见也不可执行。",

    ("后台管理系统", "工作台", "数据看板"):
        "工作台首页数据看板。顶部展示订单数、GMV、用户数等关键指标卡片；"
        "中部提供销售趋势折线图、渠道分布饼图等可视化组件；数据支持"
        "实时刷新；同时聚合待办事项列表，便于管理员快速进入待处理任务。",

    ("后台管理系统", "商品管理", "商品列表"):
        "商品列表页。提供多条件筛选（分类/状态/价格等）能力，支持商品批量"
        "上下架操作，可直接在列表行内编辑价格与库存；同时支持商品数据"
        "一键导出为 Excel 文件，方便线下运营分析。",

    ("后台管理系统", "商品管理", "商品编辑"):
        "商品编辑页。包含基本信息表单、SKU 规格管理（支持多维规格组合）、"
        "商品图片上传（多图、拖拽排序）、富文本详情编辑等功能；同时支持"
        "商品与分类、标签、营销活动的灵活关联配置。",

    ("后台管理系统", "商品管理", "商品分类"):
        "商品分类管理。采用树形结构展示多级分类，支持新增、编辑、删除与"
        "拖拽排序；同时可对单个分类进行启用 / 禁用切换，控制前台是否可见。",

    ("后台管理系统", "订单管理", "订单列表"):
        "后台订单列表页。支持按状态、时间、渠道、支付方式等多维度筛选，"
        "可查看订单详情；并提供订单导出与打印能力，支持手动改价与"
        "关闭订单等运营干预操作。",

    ("后台管理系统", "订单管理", "退款审核"):
        "退款审核模块。集中展示待审核退款单，支持同意 / 拒绝审核操作，"
        "通过退款单状态机驱动整个退款流程的流转；并提供凭证查看、"
        "管理员备注、退款打款回执登记等完整审计能力。",

    ("后台管理系统", "渠道供货", "渠道商管理"):
        "渠道商管理。提供渠道商列表与筛选能力，支持入驻审核流程；可为"
        "通过的渠道商生成专属推广二维码，并展示业绩统计与佣金预览，"
        "便于运营侧评估渠道贡献。",

    ("后台管理系统", "渠道供货", "供货商管理"):
        "供货商管理。提供供货商列表与筛选，支持资质审核流程；可配置"
        "供货范围（覆盖区域、可供商品）；并支持接单与发货等日常履约操作，"
        "完成供货履约闭环。",

    ("后台管理系统", "地区营销", "地区管理"):
        "地区管理。提供省 / 市 / 区三级联动选择能力，并支持为指定地区"
        "分配专属运营账号，实现区域化精细运营。",

    ("后台管理系统", "地区营销", "Banner 营销"):
        "Banner 营销配置。提供 Banner 列表与筛选，支持新增与编辑；可配置"
        "上下线时间、点击跳转链接与活动绑定，并提供实时预览能力，便于"
        "运营调整后即时查看效果。",

    ("后台管理系统", "财务管理", "账单"):
        "财务管理 - 账单模块。提供渠道商佣金账单展示与明细查询，"
        "并支持对账与结算状态管理，确保资金链路清晰可追溯。",

    ("后台管理系统", "财务管理", "资金流水"):
        "财务管理 - 资金流水。提供收支明细查询能力，并支持提现申请"
        "与审批流程，保障资金安全合规。",

    ("后台管理系统", "系统配置", "全局参数"):
        "系统全局参数配置。包括基础业务参数（如开关、版本号）以及"
        "运费 / 起送价等价格规则配置，作为系统运行的全局基准。",

    ("后台管理系统", "系统配置", "C 端banner修改"):
        "C 端内容装修 - Banner 配置。支持配置小程序首页 Banner 的主副标题，"
        "并提供背景图上传能力（自动 Base64 编码存储），实现运营侧对"
        "C 端 Banner 的快速调整。",

    ("后台管理系统", "系统配置", "地区运营账号"):
        "地区运营账号管理。提供账号列表查询，支持为不同账号分配"
        "对应的角色与权限范围，实现按区域、按职能的精细化权限管理。",
}


# ---------- 样式 ----------
FONT_NAME = "Microsoft YaHei"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
F_HDR = Font(name=FONT_NAME, bold=True, color="1F4E78", size=11)
F_L1 = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
F_BODY = Font(name=FONT_NAME, color="000000", size=10)
F_TOTAL = Font(name=FONT_NAME, bold=True, color="9C0006", size=11)
F_BLUE = Font(name=FONT_NAME, color="0000FF", size=10)

FILL_TITLE = PatternFill("solid", start_color="1F4E78")
FILL_HEADER = PatternFill("solid", start_color="D9E1F2")
FILL_L1 = PatternFill("solid", start_color="305496")
FILL_L2 = PatternFill("solid", start_color="8EA9DB")
FILL_TOTAL = PatternFill("solid", start_color="FCE4D6")
FILL_DESC = PatternFill("solid", start_color="FFFFFF")
FILL_BLUE = PatternFill("solid", start_color="E7E6E6")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_DESC = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)


def fill_merged(ws):
    """返回二维数组，把合并单元格内被覆盖的 None/空值用上方最近非空值补齐。"""
    rows_data = []
    last = {}  # last[c] = 该列上一次出现的非空值
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                v = last.get(c, "")
            else:
                last[c] = v
            row.append(v)
        rows_data.append(row)
    return rows_data


def collect_groups(src_path):
    wb = load_workbook(src_path)
    ws = wb.active
    data = fill_merged(ws)

    rows = []
    for r in range(2, len(data)):  # 跳过表头，从 R3 开始
        row = data[r]
        if not row or all(v == "" for v in row):
            continue
        code = row[0]
        if isinstance(code, str) and code == "合计":
            continue
        main = row[1] or ""
        l1 = row[2] or ""
        l2 = row[3] or ""
        l3 = row[4] or ""
        days = row[5] or 0
        price = row[6] or 0
        amount_cell = row[7]
        if isinstance(amount_cell, str) and amount_cell.startswith("="):
            amount = (days or 0) * (price or 0)
        else:
            amount = amount_cell or 0
        rows.append((code, main, l1, l2, l3, days, price, amount))

    # 按 (主, 一, 二) 分组
    groups = {}
    order = []
    for code, main, l1, l2, l3, days, price, amount in rows:
        key = (main, l1, l2)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append((l3, days, price, amount))
    return order, groups


def write_output(dst_path, order, groups):
    wb = Workbook()
    ws = wb.active
    ws.title = "模块清单（合并版）"

    # 列宽
    widths = [8, 14, 14, 16, 60, 8, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 标题
    last_col = len(widths)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "体彩衍生品商城 — 模块清单（二级整合版）"
    ws["A1"].font = F_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = ALIGN_C
    ws.row_dimensions[1].height = 32

    # 表头
    headers = ["编号", "主模块", "一级模块", "二级模块", "功能说明", "人天", "单价(元/天)", "金额(元)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = F_HDR
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER
    ws.row_dimensions[2].height = 26

    # 数据行
    row_idx = 3
    seq = 0
    for key in order:
        main, l1, l2 = key
        items = groups[key]
        seq += 1
        # 人天/金额汇总
        total_days = sum((it[1] or 0) for it in items)
        total_amount = sum((it[3] or 0) for it in items)
        # 取第一个出现的单价（文件中各行单价一致）
        price = next((it[2] for it in items if it[2]), 0)
        # 功能说明
        desc = DESCRIPTIONS.get(key)
        if not desc:
            # 兜底：用三级模块拼接
            desc = "；".join(it[0] for it in items if it[0]) + "。"

        ws.cell(row=row_idx, column=1, value=seq)
        ws.cell(row=row_idx, column=2, value=main)
        ws.cell(row=row_idx, column=3, value=l1)
        ws.cell(row=row_idx, column=4, value=l2)
        ws.cell(row=row_idx, column=5, value=desc)
        ws.cell(row=row_idx, column=6, value=round(total_days, 1))
        ws.cell(row=row_idx, column=7, value=price)
        ws.cell(row=row_idx, column=8, value=f"=F{row_idx}*G{row_idx}")

        # 样式
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            if c == 1:
                cell.font = F_BODY
                cell.alignment = ALIGN_C
            elif c == 2:
                cell.font = F_BODY
                cell.alignment = ALIGN_C
            elif c == 3:
                cell.font = F_L1
                cell.alignment = ALIGN_C
                cell.fill = FILL_L1
                cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
            elif c == 4:
                cell.font = F_BODY
                cell.alignment = ALIGN_C
                cell.fill = FILL_L2
                cell.font = Font(name=FONT_NAME, bold=True, color="1F4E78", size=10)
            elif c == 5:
                cell.font = F_BODY
                cell.alignment = ALIGN_DESC
                cell.fill = FILL_DESC
            elif c == 6:
                cell.font = F_BLUE
                cell.fill = FILL_BLUE
                cell.alignment = ALIGN_R
                cell.number_format = "0.0;-0.0;-"
            elif c == 7:
                cell.font = F_BLUE
                cell.fill = FILL_BLUE
                cell.alignment = ALIGN_R
                cell.number_format = "#,##0;-#,##0;-"
            elif c == 8:
                cell.font = F_BODY
                cell.alignment = ALIGN_R
                cell.number_format = "#,##0;-#,##0;-"

        # 行高按文字长度估算
        line_len = max(40, len(desc) // 35 + 1)
        ws.row_dimensions[row_idx].height = 18 * line_len + 4
        row_idx += 1

    # 合计行
    last_data_r = row_idx - 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    ws.cell(row=row_idx, column=1, value="合计")
    ws.cell(row=row_idx, column=6, value=f"=SUM(F3:F{last_data_r})")
    ws.cell(row=row_idx, column=7, value=f"=IFERROR(H{row_idx}/F{row_idx},0)")
    ws.cell(row=row_idx, column=8, value=f"=SUM(H3:H{last_data_r})")
    for c in range(1, 9):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = F_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER
        if c == 1:
            cell.alignment = ALIGN_C
        elif c == 6:
            cell.alignment = ALIGN_R
            cell.number_format = "0.0;-0.0;-"
        elif c == 7:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        elif c == 8:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
    ws.row_dimensions[row_idx].height = 28

    ws.freeze_panes = "B3"
    wb.save(dst_path)


if __name__ == "__main__":
    order, groups = collect_groups(SRC)
    write_output(DST, order, groups)
    print(f"Saved: {DST}")
    print(f"二级模块数: {len(order)}")
    print(f"汇总人天: {sum(sum(it[1] or 0 for it in groups[k]) for k in order):.1f}")
    print(f"汇总金额: {sum(sum(it[3] or 0 for it in groups[k]) for k in order):,.0f}")