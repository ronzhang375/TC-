"""Generate 3-level quotation Excel for 体彩衍生品商城小程序.

- Sheet 1: 报价汇总（C 端 + 后台 + 合计）
- Sheet 2: C 端小程序 — 3 级模块清单
- Sheet 3: 后台管理系统 — 3 级模块清单
- Sheet 4: 报价说明

Pricing: formula-driven. 单价 by VLOOKUP; 金额 = 人天 × 单价;
各一级 / 二级小计用 SUMIF 自动汇总。蓝色单元格为人天/单价/税率可调。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 角色单价（元 / 人天，2026 中国中端市场） ----------
ROLES = [
    ("前端开发", 1300),
    ("小程序开发", 1300),
    ("UI 设计师", 1000),
    ("测试工程师", 900),
]
ROLE_LOOKUP_RANGE = "角色表"

# ---------- 通用样式 ----------
FONT_NAME = "Microsoft YaHei"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 配色
FILL_TITLE = PatternFill("solid", start_color="1F4E78")     # 深蓝
FILL_L1 = PatternFill("solid", start_color="305496")        # 中蓝
FILL_L2 = PatternFill("solid", start_color="8EA9DB")        # 浅蓝
FILL_L3 = PatternFill("solid", start_color="FFFFFF")        # 白
FILL_TOTAL = PatternFill("solid", start_color="FCE4D6")     # 橙色
FILL_GRAND = PatternFill("solid", start_color="C00000")     # 深红
FILL_ROLE = PatternFill("solid", start_color="E7E6E6")      # 灰
FILL_HEADER = PatternFill("solid", start_color="D9E1F2")    # 蓝灰

# 字体
F_TITLE = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
F_L1 = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
F_L2 = Font(name=FONT_NAME, bold=True, color="1F4E78", size=11)
F_L3 = Font(name=FONT_NAME, color="000000", size=10)
F_HDR = Font(name=FONT_NAME, bold=True, color="1F4E78", size=10)
F_TOTAL = Font(name=FONT_NAME, bold=True, color="9C0006", size=11)
F_GRAND = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
F_BLUE = Font(name=FONT_NAME, color="0000FF", size=10)  # 可调输入

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def style_row(ws, row, start_col, end_col, font, fill, alignment, border=True):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        if border:
            cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ======================================================================
# C 端小程序 3 级模块数据
# ======================================================================
# 每条: (编号, 一级, 二级, 三级, 角色, 人天)
C_END = [
    # A 首页 4.0
    ("A1", "A. 首页", "Banner 轮播", "自动/手动切换",        "前端开发", 0.4),
    ("A2", "A. 首页", "Banner 轮播", "点击跳转详情/活动页",  "前端开发", 0.3),
    ("A3", "A. 首页", "Banner 轮播", "营销活动配置拉取",      "前端开发", 0.3),
    ("A4", "A. 首页", "分类入口",   "4-8 个宫格布局",          "前端开发", 0.3),
    ("A5", "A. 首页", "分类入口",   "图标与名称渲染",          "前端开发", 0.3),
    ("A6", "A. 首页", "分类入口",   "点击跳转商品列表",        "前端开发", 0.2),
    ("A7", "A. 首页", "热门商品",   "商品卡片信息渲染",        "前端开发", 0.5),
    ("A8", "A. 首页", "热门商品",   "加载更多/分页",           "前端开发", 0.5),
    ("A9", "A. 首页", "热门商品",   "快捷加入购物车",          "前端开发", 0.5),
    ("A10","A. 首页", "Tabbar 导航","4 个主 Tab 切换",         "前端开发", 0.3),
    ("A11","A. 首页", "Tabbar 导航","未读消息小红点",          "前端开发", 0.3),

    # B 商品 7.0
    ("B1", "B. 商品", "分类页",     "左侧一级分类",            "前端开发", 0.5),
    ("B2", "B. 商品", "分类页",     "右侧二级分类",            "前端开发", 0.5),
    ("B3", "B. 商品", "分类页",     "切换刷新商品",            "前端开发", 0.5),
    ("B4", "B. 商品", "商品列表",   "列表/网格视图切换",       "前端开发", 0.5),
    ("B5", "B. 商品", "商品列表",   "价格/销量/综合排序",      "前端开发", 0.8),
    ("B6", "B. 商品", "商品列表",   "多维筛选（分类/标签/价格区间）", "前端开发", 0.7),
    ("B7", "B. 商品", "商品详情",   "商品图片轮播",            "前端开发", 0.4),
    ("B8", "B. 商品", "商品详情",   "SKU 规格选择弹层",        "前端开发", 1.0),
    ("B9", "B. 商品", "商品详情",   "数量加减与库存校验",      "前端开发", 0.5),
    ("B10","B. 商品", "商品详情",   "富文本详情渲染",          "前端开发", 1.0),
    ("B11","B. 商品", "商品详情",   "加入购物车 / 立即购买",   "前端开发", 0.6),

    # C 购物车 3.0
    ("C1", "C. 购物车", "购物车列表","按店铺/供货商分组",     "前端开发", 0.5),
    ("C2", "C. 购物车", "购物车列表","单选 / 全选",            "前端开发", 0.5),
    ("C3", "C. 购物车", "购物车列表","数量增减与失效清理",     "前端开发", 0.5),
    ("C4", "C. 购物车", "购物车结算","选中商品总价合计",       "前端开发", 0.4),
    ("C5", "C. 购物车", "购物车结算","结算栏置底固定",         "前端开发", 0.4),
    ("C6", "C. 购物车", "购物车结算","去结算跳转订单确认",     "前端开发", 0.7),

    # D 订单 8.0
    ("D1", "D. 订单", "订单确认",   "默认/选择收货地址",       "前端开发", 0.5),
    ("D2", "D. 订单", "订单确认",   "商品清单与价格",          "前端开发", 0.4),
    ("D3", "D. 订单", "订单确认",   "配送方式 / 优惠券",       "前端开发", 0.5),
    ("D4", "D. 订单", "订单确认",   "费用明细与提交订单",      "前端开发", 0.6),
    ("D5", "D. 订单", "支付页",     "微信支付唤起",            "前端开发", 0.5),
    ("D6", "D. 订单", "支付页",     "支付结果轮询",            "前端开发", 0.5),
    ("D7", "D. 订单", "支付页",     "支付失败/取消处理",       "前端开发", 0.3),
    ("D8", "D. 订单", "订单列表",   "状态 Tab 切换",           "前端开发", 0.4),
    ("D9", "D. 订单", "订单列表",   "订单卡片信息",            "前端开发", 0.4),
    ("D10","D. 订单", "订单列表",   "订单操作（取消/付款/退款/确认收货）", "前端开发", 0.6),
    ("D11","D. 订单", "订单详情",   "状态流转时间线",          "前端开发", 0.4),
    ("D12","D. 订单", "订单详情",   "物流轨迹",                "前端开发", 0.4),
    ("D13","D. 订单", "订单详情",   "售后入口",                "前端开发", 0.3),
    ("D14","D. 订单", "退款申请",   "退款原因选择",            "前端开发", 0.4),
    ("D15","D. 订单", "退款申请",   "退款金额与凭证上传",      "前端开发", 0.5),
    ("D16","D. 订单", "退款申请",   "退款进度查询",            "前端开发", 0.4),
    ("D17","D. 订单", "退款申请",   "退款状态机 UI",           "前端开发", 0.4),

    # E 用户 6.0
    ("E1", "E. 用户", "登录/注册",  "微信一键登录",            "前端开发", 0.8),
    ("E2", "E. 用户", "登录/注册",  "手机号验证码登录",        "前端开发", 0.7),
    ("E3", "E. 用户", "登录/注册",  "Token 持久化与刷新",      "前端开发", 0.5),
    ("E4", "E. 用户", "个人中心",   "用户信息卡片",            "前端开发", 0.5),
    ("E5", "E. 用户", "个人中心",   "订单/收藏/优惠券入口",    "前端开发", 0.5),
    ("E6", "E. 用户", "个人中心",   "设置与退出登录",          "前端开发", 0.3),
    ("E7", "E. 用户", "收货地址",   "地址列表",                "前端开发", 0.4),
    ("E8", "E. 用户", "收货地址",   "新增 / 编辑地址",         "前端开发", 0.6),
    ("E9", "E. 用户", "收货地址",   "默认地址设置",            "前端开发", 0.3),
    ("E10","E. 用户", "收货地址",   "省市区地址选择器",        "前端开发", 0.7),
    ("E11","E. 用户", "收货地址",   "地址簿导入与校验",        "前端开发", 0.4),

    # F 公共能力 10.0
    ("F1", "F. 公共能力", "网络请求", "Axios / uni.request 封装", "前端开发", 1.0),
    ("F2", "F. 公共能力", "网络请求", "请求 / 响应拦截器",     "前端开发", 0.8),
    ("F3", "F. 公共能力", "网络请求", "业务模块 API 拆分（user/product/order/cart/pay/refund 等）", "前端开发", 1.2),
    ("F4", "F. 公共能力", "状态管理", "Pinia Store 搭建",       "前端开发", 0.8),
    ("F5", "F. 公共能力", "状态管理", "用户 / 购物车 / 系统 Store", "前端开发", 1.5),
    ("F6", "F. 公共能力", "状态管理", "数据持久化",            "前端开发", 0.5),
    ("F7", "F. 公共能力", "UI 组件", "弹窗 / Toast / Loading", "前端开发", 1.2),
    ("F8", "F. 公共能力", "UI 组件", "业务组件（Empty/Error/Price/Countdown）", "前端开发", 1.5),
    ("F9", "F. 公共能力", "UI 组件", "主题与配色（奶油玫瑰粉）", "UI 设计师", 0.5),
    ("F10","F. 公共能力", "公共 Shell","路由配置",              "前端开发", 0.4),
    ("F11","F. 公共能力", "公共 Shell","登录守卫与权限拦截",    "前端开发", 0.3),
    ("F12","F. 公共能力", "公共 Shell","错误页 / 空状态 / 离线", "前端开发", 0.3),
]

# ======================================================================
# 后台管理 3 级模块数据
# ======================================================================
ADMIN = [
    # G 系统认证 3.0
    ("G1", "G. 系统认证", "登录",   "账号密码登录",            "前端开发", 0.6),
    ("G2", "G. 系统认证", "登录",   "图形验证码",              "前端开发", 0.4),
    ("G3", "G. 系统认证", "登录",   "记住登录 / 自动续签",     "前端开发", 0.3),
    ("G4", "G. 系统认证", "RBAC 权限","角色与菜单管理",         "前端开发", 0.6),
    ("G5", "G. 系统认证", "RBAC 权限","动态菜单渲染",           "前端开发", 0.5),
    ("G6", "G. 系统认证", "RBAC 权限","按钮级操作权限",         "前端开发", 0.6),

    # H 工作台 3.0
    ("H1", "H. 工作台", "数据看板", "关键指标卡片（订单/GMV/用户）", "前端开发", 0.7),
    ("H2", "H. 工作台", "数据看板", "销售趋势折线图",          "前端开发", 0.7),
    ("H3", "H. 工作台", "数据看板", "渠道分布饼图",            "前端开发", 0.6),
    ("H4", "H. 工作台", "数据看板", "实时数据刷新",            "前端开发", 0.5),
    ("H5", "H. 工作台", "数据看板", "待办事项列表",            "前端开发", 0.5),

    # I 商品管理 6.0
    ("I1", "I. 商品管理", "商品列表","多条件筛选",              "前端开发", 0.5),
    ("I2", "I. 商品管理", "商品列表","批量上下架",              "前端开发", 0.5),
    ("I3", "I. 商品管理", "商品列表","行内价格/库存编辑",       "前端开发", 0.5),
    ("I4", "I. 商品管理", "商品列表","商品导出 Excel",          "前端开发", 0.4),
    ("I5", "I. 商品管理", "商品编辑","基本信息表单",            "前端开发", 0.8),
    ("I6", "I. 商品管理", "商品编辑","SKU 规格管理（多维规格组合）", "前端开发", 1.2),
    ("I7", "I. 商品管理", "商品编辑","图片上传（多图/拖拽）",   "前端开发", 0.5),
    ("I8", "I. 商品管理", "商品编辑","富文本详情编辑",          "前端开发", 0.7),
    ("I9", "I. 商品管理", "商品编辑","商品关联（分类/标签/活动）", "前端开发", 0.4),
    ("I10","I. 商品管理", "商品分类","树形结构渲染",            "前端开发", 0.5),
    ("I11","I. 商品管理", "商品分类","增删改与拖拽排序",        "前端开发", 0.4),
    ("I12","I. 商品管理", "商品分类","启用 / 禁用切换",         "前端开发", 0.4),

    # J 订单管理 4.0
    ("J1", "J. 订单管理", "订单列表","多维筛选（状态/时间/渠道/支付）", "前端开发", 0.6),
    ("J2", "J. 订单管理", "订单列表","订单详情查看",            "前端开发", 0.6),
    ("J3", "J. 订单管理", "订单列表","订单导出 / 打印",        "前端开发", 0.5),
    ("J4", "J. 订单管理", "订单列表","手动改价 / 关闭订单",     "前端开发", 0.4),
    ("J5", "J. 订单管理", "退款审核","待审核列表",              "前端开发", 0.6),
    ("J6", "J. 订单管理", "退款审核","同意 / 拒绝流程",         "前端开发", 0.6),
    ("J7", "J. 订单管理", "退款审核","退款单状态机",            "前端开发", 0.6),
    ("J8", "J. 订单管理", "退款审核","凭证查看 / 备注",         "前端开发", 0.4),
    ("J9", "J. 订单管理", "退款审核","退款打款回执",            "前端开发", 0.3),

    # K 渠道商 / 供货商 4.5
    ("K1", "K. 渠道供货", "渠道商管理","列表与筛选",            "前端开发", 0.4),
    ("K2", "K. 渠道供货", "渠道商管理","入驻审核",              "前端开发", 0.7),
    ("K3", "K. 渠道供货", "渠道商管理","渠道二维码生成",        "前端开发", 0.5),
    ("K4", "K. 渠道供货", "渠道商管理","业绩统计与佣金预览",    "前端开发", 0.5),
    ("K5", "K. 渠道供货", "供货商管理","列表与筛选",            "前端开发", 0.4),
    ("K6", "K. 渠道供货", "供货商管理","资质审核",              "前端开发", 0.7),
    ("K7", "K. 渠道供货", "供货商管理","供货范围配置",          "前端开发", 0.5),
    ("K8", "K. 渠道供货", "供货商管理","接单 / 发货操作",       "前端开发", 0.8),

    # L 地区与营销 3.0
    ("L1", "L. 地区营销", "地区管理","省市区三级联动",          "前端开发", 0.5),
    ("L2", "L. 地区营销", "地区管理","地区运营账号分配",        "前端开发", 0.5),
    ("L3", "L. 地区营销", "Banner 营销","列表与筛选",           "前端开发", 0.4),
    ("L4", "L. 地区营销", "Banner 营销","新增 / 编辑",          "前端开发", 0.5),
    ("L5", "L. 地区营销", "Banner 营销","上下线时间配置",       "前端开发", 0.3),
    ("L6", "L. 地区营销", "Banner 营销","点击跳转 / 活动绑定",   "前端开发", 0.4),
    ("L7", "L. 地区营销", "Banner 营销","实时预览",             "前端开发", 0.4),

    # M 财务 2.0
    ("M1", "M. 财务管理", "账单",   "渠道商佣金账单",          "前端开发", 0.5),
    ("M2", "M. 财务管理", "账单",   "账单明细",                "前端开发", 0.4),
    ("M3", "M. 财务管理", "账单",   "对账与结算状态",          "前端开发", 0.4),
    ("M4", "M. 财务管理", "资金流水","收支明细查询",            "前端开发", 0.4),
    ("M5", "M. 财务管理", "资金流水","提现申请与审批",          "前端开发", 0.3),

    # N 系统配置 3.5
    ("N1", "N. 系统配置", "全局参数","基础参数配置",            "前端开发", 0.4),
    ("N2", "N. 系统配置", "全局参数","运费 / 起送价规则",       "前端开发", 0.4),
    ("N3", "N. 系统配置", "C 端内容装修","Banner 主副标题",     "前端开发", 0.3),
    ("N4", "N. 系统配置", "C 端内容装修","背景图上传 Base64",   "前端开发", 0.4),
    ("N5", "N. 系统配置", "C 端内容装修","所见即所得预览",      "前端开发", 0.4),
    ("N6", "N. 系统配置", "C 端内容装修","双端 storage 实时同步", "前端开发", 0.8),
    ("N7", "N. 系统配置", "C 端内容装修","装修主题视觉设计",    "UI 设计师", 0.5),
    ("N8", "N. 系统配置", "地区运营账号","账号列表",            "前端开发", 0.4),
    ("N9", "N. 系统配置", "地区运营账号","角色与权限分配",      "前端开发", 0.4),

    # O 公共 Shell 3.0
    ("O1", "O. 公共 Shell", "框架",   "路由 / 菜单配置",        "前端开发", 0.5),
    ("O2", "O. 公共 Shell", "框架",   "主布局 / 面包屑 / 顶栏", "前端开发", 0.5),
    ("O3", "O. 公共 Shell", "框架",   "主题样式（Element Plus）", "UI 设计师", 0.4),
    ("O4", "O. 公共 Shell", "通用能力","表格 / 表单 / 分页组件", "前端开发", 0.6),
    ("O5", "O. 公共 Shell", "通用能力","上传 / 导出工具",       "前端开发", 0.5),
    ("O6", "O. 公共 Shell", "通用能力","全局错误处理 / 日志",    "前端开发", 0.5),
]


# ======================================================================
# 工作表 1: 报价汇总
# ======================================================================
def build_summary(wb, c_end_rows, admin_rows):
    ws = wb.active
    ws.title = "1.报价汇总"
    set_col_widths(ws, [4, 28, 14, 16, 14, 16, 14, 16])

    # 标题
    ws.merge_cells("B2:H2")
    ws["B2"] = "体彩衍生品商城小程序 — 系统报价汇总"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = ALIGN_C
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:H3")
    ws["B3"] = "范围：C 端小程序 + 运营管理后台（功能点细化到三级模块）"
    ws["B3"].font = Font(name=FONT_NAME, italic=True, color="595959", size=10)
    ws["B3"].alignment = ALIGN_C
    ws.row_dimensions[3].height = 18

    # 表头：9 列布局 (B-J)
    # B: C 端一级 / C: C 端人天 / D: C 端金额
    # E: 后台一级 / F: 后台人天 / G: 后台金额
    # H: 合计人天 / I: 合计金额
    headers = [
        "C 端一级模块", "人天", "金额（元）",
        "后台一级模块", "人天", "金额（元）",
        "合计人天", "合计金额（元）",
    ]
    for i, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = F_HDR
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = BORDER
    ws.row_dimensions[5].height = 28

    # 7 个一级模块（按顺序），关联 C 端和后台
    l1_modules = [
        ("A. 首页",     "A. 首页",     "2.C端模块", "A"),
        ("B. 商品",     "I. 商品管理", "2.C端模块", "B"),
        ("C. 购物车",   "—",            "2.C端模块", "C"),
        ("D. 订单",     "J. 订单管理", "3.后台模块", "D"),
        ("E. 用户",     "G. 系统认证", "2.C端模块", "E"),
        ("F. 公共能力", "L. 地区营销", "3.后台模块", "F"),
        ("—",            "M. 财务管理", "3.后台模块", None),
        ("—",            "N. 系统配置", "3.后台模块", None),
        ("—",            "O. 公共 Shell","3.后台模块", None),
    ]
    # 由于两侧一级模块并不一一对应（合并汇总会重复），改用更简单的方式：
    # 左侧仅列 C 端 6 个一级，右侧列后台 9 个一级，末列对每一行求和
    # 先清空上面预定义，改用扁平化：每行展示一个一级模块
    l1_modules = [
        # (C端一级, 后台一级)
        ("A. 首页",      "H. 工作台"),
        ("B. 商品",      "I. 商品管理"),
        ("C. 购物车",    "J. 订单管理"),
        ("D. 订单",      "K. 渠道供货"),
        ("E. 用户",      "L. 地区营销"),
        ("F. 公共能力",  "M. 财务管理"),
        ("",             "G. 系统认证"),
        ("",             "N. 系统配置"),
        ("",             "O. 公共 Shell"),
    ]

    for i, (c_l1, a_l1) in enumerate(l1_modules):
        r = 6 + i
        ws.cell(row=r, column=2, value=c_l1)
        if c_l1:
            ws.cell(row=r, column=3,
                    value=f"=SUMIF('2.C端模块'!C:C,B{r},'2.C端模块'!G:G)")
            ws.cell(row=r, column=4,
                    value=f"=SUMIF('2.C端模块'!C:C,B{r},'2.C端模块'!I:I)")
        else:
            ws.cell(row=r, column=3, value="")
            ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5,
                value=f"=SUMIF('3.后台模块'!C:C,F{r},'3.后台模块'!G:G)" if a_l1 else "")
        ws.cell(row=r, column=6,
                value=f"=SUMIF('3.后台模块'!C:C,F{r},'3.后台模块'!I:I)" if a_l1 else "")
        # 后台一级放在 F 列前 — 把 E 列改名 F 列
        ws.cell(row=r, column=5, value=a_l1 if a_l1 else "")
        ws.cell(row=r, column=6,
                value=f"=SUMIF('3.后台模块'!C:C,E{r},'3.后台模块'!G:G)" if a_l1 else "")
        ws.cell(row=r, column=7,
                value=f"=SUMIF('3.后台模块'!C:C,E{r},'3.后台模块'!I:I)" if a_l1 else "")

        ws.cell(row=r, column=8, value=f"=IFERROR(C{r}+F{r},0)")  # 合计人天
        ws.cell(row=r, column=9, value=f"=IFERROR(D{r}+G{r},0)")  # 合计金额

    # 实际列结构： B=一级(C)  C=人天(C)  D=金额(C)
    #                E=一级(B)  F=人天(B)  G=金额(B)
    #                H=合计人天  I=合计金额
    # 修正上面的错位
    for r in range(6, 6 + len(l1_modules)):
        ws.cell(row=r, column=2, value=l1_modules[r-6][0])  # C 端一级
        if l1_modules[r-6][0]:
            ws.cell(row=r, column=3, value=f"=SUMIF('2.C端模块'!C:C,B{r},'2.C端模块'!G:G)")
            ws.cell(row=r, column=4, value=f"=SUMIF('2.C端模块'!C:C,B{r},'2.C端模块'!I:I)")
        ws.cell(row=r, column=5, value=l1_modules[r-6][1])  # 后台一级
        if l1_modules[r-6][1]:
            ws.cell(row=r, column=6, value=f"=SUMIF('3.后台模块'!C:C,E{r},'3.后台模块'!G:G)")
            ws.cell(row=r, column=7, value=f"=SUMIF('3.后台模块'!C:C,E{r},'3.后台模块'!I:I)")
        ws.cell(row=r, column=8, value=f"=IFERROR(C{r},0)+IFERROR(F{r},0)")
        ws.cell(row=r, column=9, value=f"=IFERROR(D{r},0)+IFERROR(G{r},0)")

    # 样式
    last_l1 = 6 + len(l1_modules) - 1
    for r in range(6, last_l1 + 1):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.font = F_L3
            cell.border = BORDER
            if c in (3, 6, 8):
                cell.alignment = ALIGN_R
                cell.number_format = "0.0;-0.0;-"
            elif c in (4, 7, 9):
                cell.alignment = ALIGN_R
                cell.number_format = "#,##0;-#,##0;-"
            else:
                cell.alignment = ALIGN_L
        ws.row_dimensions[r].height = 22

    # C 端小计行
    c_total_r = last_l1 + 1
    ws.cell(row=c_total_r, column=2, value="C 端小计")
    ws.cell(row=c_total_r, column=3, value=f"=SUM(C6:C{last_l1})")
    ws.cell(row=c_total_r, column=4, value=f"=SUM(D6:D{last_l1})")
    ws.cell(row=c_total_r, column=5, value="后台小计")
    ws.cell(row=c_total_r, column=6, value=f"=SUM(F6:F{last_l1})")
    ws.cell(row=c_total_r, column=7, value=f"=SUM(G6:G{last_l1})")
    ws.cell(row=c_total_r, column=8, value=f"=C{c_total_r}+F{c_total_r}")
    ws.cell(row=c_total_r, column=9, value=f"=D{c_total_r}+G{c_total_r}")
    for c in range(2, 10):
        cell = ws.cell(row=c_total_r, column=c)
        cell.font = F_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER
        if c in (3, 6, 8):
            cell.alignment = ALIGN_R
            cell.number_format = "0.0;-0.0;-"
        elif c in (4, 7, 9):
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        else:
            cell.alignment = ALIGN_C
    ws.row_dimensions[c_total_r].height = 26

    # 角色单价表
    rate_r = c_total_r + 2
    ws.cell(row=rate_r, column=2, value="角色单价比照表（可调）")
    ws.cell(row=rate_r, column=2).font = F_L2
    ws.cell(row=rate_r, column=2).fill = FILL_L2
    ws.merge_cells(start_row=rate_r, start_column=2, end_row=rate_r, end_column=4)
    ws.cell(row=rate_r, column=2).alignment = ALIGN_C
    ws.row_dimensions[rate_r].height = 22

    for i, h in enumerate(["角色", "元/人天", "说明"]):
        cell = ws.cell(row=rate_r + 1, column=2 + i, value=h)
        cell.font = F_HDR
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = BORDER
    for i, (role, rate) in enumerate(ROLES):
        r = rate_r + 2 + i
        ws.cell(row=r, column=2, value=role)
        ws.cell(row=r, column=3, value=rate)
        ws.cell(row=r, column=3).font = F_BLUE  # 蓝色可调
        ws.cell(row=r, column=3).fill = FILL_ROLE
        ws.cell(row=r, column=4, value={
            "前端开发": "含 Web 端与管理后台页面开发",
            "小程序开发": "含 UniApp 跨端页面与组件",
            "UI 设计师": "含视觉升级、主题与装修",
            "测试工程师": "含功能、回归与上线验收",
        }.get(role, ""))
        for c in range(2, 5):
            cell = ws.cell(row=r, column=c)
            if c != 3:
                cell.font = F_L3
            cell.border = BORDER
            if c == 3:
                cell.alignment = ALIGN_R
                cell.number_format = "#,##0;-#,##0;-"
            else:
                cell.alignment = ALIGN_L

    # 总价行
    grand_r = c_total_r + 9
    ws.cell(row=grand_r, column=2, value="开发费用合计（C 端 + 后台）")
    ws.merge_cells(start_row=grand_r, start_column=2, end_row=grand_r, end_column=7)
    ws.cell(row=grand_r, column=8, value=f"=H{c_total_r}")
    ws.cell(row=grand_r, column=9, value=f"=I{c_total_r}")
    for c in range(2, 10):
        cell = ws.cell(row=grand_r, column=c)
        cell.font = F_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER
        if c == 8:
            cell.alignment = ALIGN_R
            cell.number_format = "0.0;-0.0;-"
        elif c == 9:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        else:
            cell.alignment = ALIGN_C
    ws.row_dimensions[grand_r].height = 26

    # 税率与含税
    ws.cell(row=grand_r + 1, column=2, value="税率（增值税）")
    ws.merge_cells(start_row=grand_r + 1, start_column=2, end_row=grand_r + 1, end_column=7)
    ws.cell(row=grand_r + 1, column=8, value=0.13)
    ws.cell(row=grand_r + 1, column=8).font = F_BLUE
    ws.cell(row=grand_r + 1, column=8).fill = FILL_ROLE
    ws.cell(row=grand_r + 1, column=8).number_format = "0.0%"
    ws.cell(row=grand_r + 1, column=9, value=f"=I{grand_r}*H{grand_r + 1}")
    for c in range(2, 10):
        cell = ws.cell(row=grand_r + 1, column=c)
        cell.border = BORDER
        if c not in (8, 9):
            cell.font = F_L3
            cell.alignment = ALIGN_C
        elif c == 8:
            cell.alignment = ALIGN_R
        else:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"

    ws.cell(row=grand_r + 2, column=2, value="项目总价（含税）")
    ws.merge_cells(start_row=grand_r + 2, start_column=2, end_row=grand_r + 2, end_column=7)
    ws.cell(row=grand_r + 2, column=9, value=f"=I{grand_r}+I{grand_r + 1}")
    for c in range(2, 10):
        cell = ws.cell(row=grand_r + 2, column=c)
        cell.font = F_GRAND
        cell.fill = FILL_GRAND
        cell.border = BORDER
        if c == 9:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        else:
            cell.alignment = ALIGN_C
    ws.row_dimensions[grand_r + 2].height = 32

    # 备注
    note_r = grand_r + 4
    notes = [
        "备注：",
        "1. 工作量单位为「人天」，按 8 小时有效工时计算。",
        "2. 蓝色单元格为可调参数：角色单价、税率、人天；调整后下方所有金额自动重算。",
        "3. 包含：源代码、PRD 文档、数据库设计、测试用例、部署文档、培训与上线支持。",
        "4. 不包含：服务器 / 短信 / 支付通道 / 第三方接口年费、设计素材版权采购。",
        "5. 付款方式：30% 启动 + 40% 中期验收 + 25% 上线 + 5% 质保（3 个月）。",
        "6. 本报价有效期 30 天，最终以双方合同为准。",
    ]
    for i, t in enumerate(notes):
        cell = ws.cell(row=note_r + i, column=2, value=t)
        cell.font = F_TITLE if i == 0 else F_L3
        if i == 0:
            cell.font = Font(name=FONT_NAME, bold=True, color="1F4E78", size=10)
        ws.merge_cells(start_row=note_r + i, start_column=2,
                       end_row=note_r + i, end_column=9)
        cell.alignment = ALIGN_L

    ws.freeze_panes = "B6"


# ======================================================================
# 工作表 2/3: 3 级模块清单
# ======================================================================
HEADERS_L3 = [
    "编号",   # A
    "层级",   # B
    "一级模块",  # C
    "二级模块",  # D
    "三级模块",  # E
    "角色",   # F
    "人天",   # G
    "单价（元/天）",  # H
    "金额（元）",    # I
]


def build_module_sheet(wb, sheet_name, rows, role_range):
    ws = wb.create_sheet(sheet_name)
    set_col_widths(ws, [8, 6, 14, 18, 42, 14, 8, 14, 14])

    # 标题
    last_col = len(HEADERS_L3)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = f"体彩衍生品商城 — {sheet_name.split('.')[1]} 3 级模块清单"
    ws["A1"].font = F_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = ALIGN_C
    ws.row_dimensions[1].height = 30

    # 表头
    for i, h in enumerate(HEADERS_L3, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = F_HDR
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = BORDER
    ws.row_dimensions[3].height = 28

    # 角色单价表（F:H 列底部）
    role_anchor_r = 4 + len(rows) + 2  # 留空
    ws.cell(row=role_anchor_r, column=5, value="角色单价比照表").font = F_L2
    ws.cell(row=role_anchor_r, column=5).fill = FILL_L2
    ws.merge_cells(start_row=role_anchor_r, start_column=5,
                   end_row=role_anchor_r, end_column=6)
    ws.cell(row=role_anchor_r, column=5).alignment = ALIGN_C
    for i, h in enumerate(["角色", "元/人天"]):
        cell = ws.cell(row=role_anchor_r + 1, column=5 + i, value=h)
        cell.font = F_HDR
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = BORDER
    for i, (role, rate) in enumerate(ROLES):
        rr = role_anchor_r + 2 + i
        ws.cell(row=rr, column=5, value=role).font = F_L3
        ws.cell(row=rr, column=6, value=rate).font = F_BLUE
        ws.cell(row=rr, column=6).fill = FILL_ROLE
        ws.cell(row=rr, column=6).number_format = "#,##0;-#,##0;-"
        ws.cell(row=rr, column=6).alignment = ALIGN_R
        for c in (5, 6):
            ws.cell(row=rr, column=c).border = BORDER
    role_last_r = role_anchor_r + 1 + len(ROLES)
    # 命名范围（用于 VLOOKUP）
    range_ref = f"'{sheet_name}'!$E${role_anchor_r + 2}:$F${role_last_r}"

    # 数据行
    start_data_r = 4
    for i, (code, l1, l2, l3, role, days) in enumerate(rows):
        r = start_data_r + i
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=3)        # 层级：3
        ws.cell(row=r, column=3, value=l1)
        ws.cell(row=r, column=4, value=l2)
        ws.cell(row=r, column=5, value=l3)
        ws.cell(row=r, column=6, value=role)
        ws.cell(row=r, column=7, value=days)
        ws.cell(row=r, column=7).font = F_BLUE
        ws.cell(row=r, column=7).fill = FILL_ROLE
        ws.cell(row=r, column=7).number_format = "0.0;-0.0;-"
        # 单价 VLOOKUP
        ws.cell(row=r, column=8, value=f"=VLOOKUP(F{r},{range_ref},2,FALSE)")
        ws.cell(row=r, column=8).number_format = "#,##0;-#,##0;-"
        # 金额 = 人天 × 单价
        ws.cell(row=r, column=9, value=f"=G{r}*H{r}")
        ws.cell(row=r, column=9).number_format = "#,##0;-#,##0;-"

        # 样式
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c == 1:
                cell.alignment = ALIGN_C
                cell.font = F_L3
            elif c == 2:
                cell.alignment = ALIGN_C
                cell.font = F_L3
            elif c == 3:
                cell.alignment = ALIGN_L
                cell.font = F_L2
            elif c == 4:
                cell.alignment = ALIGN_L
                cell.font = F_L3
            elif c == 5:
                cell.alignment = ALIGN_L
                cell.font = F_L3
            elif c == 6:
                cell.alignment = ALIGN_C
                cell.font = F_L3
            elif c in (7, 8, 9):
                cell.alignment = ALIGN_R
        ws.row_dimensions[r].height = 18

    # 合计行
    last_data_r = start_data_r + len(rows) - 1
    total_r = last_data_r + 1
    ws.cell(row=total_r, column=1, value="合计")
    ws.merge_cells(start_row=total_r, start_column=1,
                   end_row=total_r, end_column=6)
    ws.cell(row=total_r, column=7, value=f"=SUM(G{start_data_r}:G{last_data_r})")
    ws.cell(row=total_r, column=9, value=f"=SUM(I{start_data_r}:I{last_data_r})")
    ws.cell(row=total_r, column=8, value=f"=IFERROR(I{total_r}/G{total_r},0)")
    for c in range(1, 10):
        cell = ws.cell(row=total_r, column=c)
        cell.font = F_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER
        if c == 7:
            cell.alignment = ALIGN_R
            cell.number_format = "0.0;-0.0;-"
        elif c == 8:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        elif c == 9:
            cell.alignment = ALIGN_R
            cell.number_format = "#,##0;-#,##0;-"
        else:
            cell.alignment = ALIGN_C
    ws.row_dimensions[total_r].height = 26

    ws.freeze_panes = "C4"


# ======================================================================
# 工作表 4: 报价说明
# ======================================================================
def build_notes(wb):
    ws = wb.create_sheet("4.报价说明")
    set_col_widths(ws, [4, 22, 90])

    ws.merge_cells("B2:C2")
    ws["B2"] = "体彩衍生品商城 — 报价说明"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = ALIGN_C
    ws.row_dimensions[2].height = 30

    sections = [
        ("1. 项目范围", [
            "本次报价覆盖 C 端小程序（UniApp + Vue 3 + Pinia）与运营管理后台（Vue 3 + Vite + Element Plus）的完整前端开发工作。",
            "后端服务（Spring Boot 3.2 + MyBatis Plus + MySQL + Redis）按现有代码基础进行联调与对接。",
        ]),
        ("2. 计价方式", [
            "按三级功能模块逐项核算：每个三级模块给出工时（人天），按角色单价折算金额。",
            "一级、二级小计使用 SUMIF 公式对三级模块自动汇总。",
            "总计 = C 端小计 + 后台小计；项目总价 = 总计 × (1 + 13% 增值税)。",
        ]),
        ("3. 角色单价依据", [
            "前端开发 1300 元/人天：参照 2026 年中国中端市场前端工程师外包均价。",
            "UI 设计师 1000 元/人天：参照中级 UI 设计师外包均价。",
            "单价可在「1.报价汇总」右下角或各模块表底部「角色单价比照表」中调整。",
        ]),
        ("4. 工时估算依据", [
            "依据现有代码体量、UI 模拟器复杂度、业务流程深度综合评估。",
            "C 端 13 页面 + 公共能力 ≈ 38 人天；后台 11 页面 + 公共 Shell ≈ 29 人天。",
            "含：基础页面、复杂交互（SKU 弹层、订单状态机、退款流程）、可视化配置（Banner 装修、双端同步）。",
        ]),
        ("5. 包含项", [
            "全部三级模块对应的源代码、组件库封装。",
            "与后端接口联调、Mock 数据、接口文档（Swagger 注释）。",
            "PRD、数据库设计、测试用例、部署文档（基于现有项目文档）。",
            "上线前 2 周问题修复与上线后 3 个月质保。",
        ]),
        ("6. 不包含项", [
            "服务器、域名、SSL 证书、短信通道、微信支付通道年费。",
            "第三方素材版权采购（图片、字体、图标库）。",
            "运营数据初始化、客服培训、推广物料设计。",
            "超过 3 个月的免费维护期。",
        ]),
        ("7. 付款方式", [
            "30% 合同签订后 5 个工作日内支付（启动款）。",
            "40% 中期里程碑（核心模块完成并通过内部测试）验收后支付。",
            "25% 系统上线后 5 个工作日内支付。",
            "5% 质保金：上线 3 个月无重大问题后支付。",
        ]),
        ("8. 风险与变更", [
            "需求变更：单次变更超过 5% 工作量时，按变更工作量单独核算。",
            "接口变更：后端接口字段调整超过 10 处时，按 0.5 人天/次补差。",
            "紧急加期：项目期间如需加急（< 原周期 70%），加收 20% 费用。",
        ]),
        ("9. 交付物清单", [
            "C 端小程序源代码（UniApp 跨端工程）。",
            "后台管理系统源代码（Vue 3 + Vite + Element Plus）。",
            "公共组件库、API 封装、状态管理 Store。",
            "与现有后端联调的接口对接说明。",
            "测试用例文档、部署文档、操作手册。",
        ]),
        ("10. 验收标准", [
            "全部三级模块按 PRD 实现并通过测试用例。",
            "关键流程：登录、购物流程、订单生命周期、退款流程、Banner 装修双端同步。",
            "性能：首屏 < 2s，交互响应 < 300ms，兼容主流移动设备与浏览器。",
            "代码质量：通过 ESLint 静态检查，关键模块单元测试覆盖率 ≥ 60%。",
        ]),
    ]

    r = 4
    for title, items in sections:
        cell = ws.cell(row=r, column=2, value=title)
        cell.font = F_L2
        cell.fill = FILL_L2
        cell.alignment = ALIGN_L
        cell.border = BORDER
        cell2 = ws.cell(row=r, column=3, value="")
        cell2.fill = FILL_L2
        cell2.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value="").border = BORDER
            c3 = ws.cell(row=r, column=3, value="• " + item)
            c3.alignment = ALIGN_L
            c3.font = F_L3
            c3.border = BORDER
            ws.row_dimensions[r].height = 30
            r += 1


# ======================================================================
# 主入口
# ======================================================================
def main():
    wb = Workbook()
    build_summary(wb, C_END, ADMIN)
    build_module_sheet(wb, "2.C端模块", C_END, None)
    build_module_sheet(wb, "3.后台模块", ADMIN, None)
    build_notes(wb)
    out = "体彩衍生品商城_系统报价表.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
